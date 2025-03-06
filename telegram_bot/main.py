import asyncio
import logging
import os
import sys
from decimal import Decimal

from aiogram import F
from aiogram.enums.parse_mode import ParseMode
from aiogram.types import Message, CallbackQuery, InlineKeyboardButton, InlineQuery, InlineQueryResultArticle, \
    InputTextMessageContent, LabeledPrice, PreCheckoutQuery, FSInputFile
from aiogram.filters import CommandStart, StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.methods import DeleteWebhook
from aiogram.utils.keyboard import InlineKeyboardBuilder
from asgiref.sync import sync_to_async
from django.contrib.humanize.templatetags.humanize import intcomma
from openpyxl import Workbook, load_workbook

import keyboards as KB
from config import bot, config, dp, faq_data
from states import CartStates, CheckoutStates
from utils import check_channels_subscription
from bot.models import CartItem, Category, Item, ShoppingCart, TelegramUser


async def get_cart_total(user_id: int) -> Decimal:
    try:
        cart = await sync_to_async(ShoppingCart.objects.get)(user__user_id=user_id)
        cart_items = await sync_to_async(list)(cart.cart_items.all())

        total = Decimal('0')
        for item in cart_items:
            price = await sync_to_async(lambda: item.item.price)()
            total += price * Decimal(item.quantity)

        return total.quantize(Decimal('0.00'))
    except Exception as e:
        logging.error(f"Ошибка при расчете суммы корзины: {e}")
        return Decimal('0.00')


def save_order_to_excel(user_id: int, address: str, total_amount: int, items: list):
    if not os.path.exists(config.EXCEL_FILE_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "Orders"
        ws.append(["User ID", "Address", "Total Amount", "Items"])
    else:
        wb = load_workbook(config.EXCEL_FILE_PATH)
        ws = wb.active

    ws.append([user_id, address, total_amount, ", ".join(items)])
    wb.save(config.EXCEL_FILE_PATH)


@dp.message(CommandStart())
async def command_start(message: Message):
    from_user = message.from_user
    user, is_created = await TelegramUser.async_create(
        user_id=from_user.id,
        first_name=from_user.first_name,
        last_name=from_user.last_name,
        username=from_user.username
    )
    if is_created:
        logging.info(f'!Новый пользователь! {user}')
    if await check_channels_subscription(from_user.id):
        await main_menu(CallbackQuery(
            id="dummy_id",
            from_user=from_user,
            chat_instance="dummy_chat_instance",
            message=message,
            data="menu"
        ))
    else:
        await message.answer(
            'Для работы бота необходимо подписаться на каналы.',
            reply_markup=KB.channels_links
        )


@dp.callback_query(F.data == 'menu')
async def main_menu(callback: CallbackQuery):
    await callback.message.answer(
        text=config.MAIN_MENU_TEXT,
        reply_markup=KB.menu,
        parse_mode=ParseMode.HTML
    )


@dp.callback_query(F.data == "check_subscriptions")
async def check_subscriptions(callback: CallbackQuery):
    user_id = callback.from_user.id
    await callback.answer('Проверяем подписки')
    if await check_channels_subscription(user_id):
        callback.data = 'menu'
        await main_menu(callback)
    else:
        await callback.message.answer(
            'Проверьте подписки на каналы.',
            reply_markup=KB.channels_links
        )


@dp.callback_query(F.data == "catalog")
async def show_categories(callback: CallbackQuery):
    categories = await sync_to_async(list)(Category.objects.all())

    builder = InlineKeyboardBuilder()
    for category in categories:
        builder.button(
            text=category.title,
            callback_data=f"category_{category.id}"
        )

    builder.button(text="🔙 Назад", callback_data="menu")

    builder.adjust(1)

    await callback.message.edit_text(
        text="📂 Выберите категорию:",
        reply_markup=builder.as_markup()
    )


@dp.callback_query(F.data.startswith("category_"))
async def show_category_items(callback: CallbackQuery, state: FSMContext):
    category_id = int(callback.data.split("_")[-1])
    await state.update_data(category_id=category_id, page=1)
    await display_category_items(callback, state)


async def display_category_items(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    category_id = data.get("category_id")
    page = data.get("page", 1)

    category = await sync_to_async(Category.objects.get)(id=category_id)

    items = await sync_to_async(list)(
        Item.objects.filter(
            category_id=category_id,
            stock__gt=0
        )[(page - 1) * config.ITEMS_PER_PAGE:page * config.ITEMS_PER_PAGE]
    )

    builder = InlineKeyboardBuilder()
    for item in items:
        builder.button(
            text=f"{item.title} - {item.price_view()} (В наличии: {item.stock})",
            callback_data=f"item_{item.id}"
        )

    total_items = await sync_to_async(
        Item.objects.filter(category_id=category_id).count
    )()
    total_pages = (
        total_items + config.ITEMS_PER_PAGE - 1
    ) // config.ITEMS_PER_PAGE

    pagination_buttons = []
    if page > 1:
        pagination_buttons.append(InlineKeyboardButton(text="⬅️ Назад", callback_data=f"prev_page_{page}"))
    if page < total_pages:
        pagination_buttons.append(InlineKeyboardButton(text="Вперед ➡️", callback_data=f"next_page_{page}"))

    if pagination_buttons:
        builder.row(*pagination_buttons)

    builder.row(InlineKeyboardButton(text="🔙 Назад", callback_data="catalog"))

    builder.adjust(1)

    await callback.message.edit_text(
        text=f"🛍️ Товары в категории {category.title}:",
        reply_markup=builder.as_markup()
    )


@dp.callback_query(F.data.startswith("item_"))
async def show_item_card(callback: CallbackQuery):
    item_id = int(callback.data.split("_")[-1])
    item = await sync_to_async(Item.objects.get)(id=item_id)

    builder = InlineKeyboardBuilder()
    builder.button(text="Добавить в корзину", callback_data=f"add_to_cart_{item.id}")
    builder.button(text="🔙 Назад", callback_data=f"category_{await sync_to_async(lambda: item.category.id)()}")

    if not item.image or not item.image.url:
        await callback.message.answer(
            f"📦 {item.title}\n\n"
            f"📝 <b>Описание:</b>\n{item.description[:500]}...\n"
            f"💵 <b>Цена:</b> {item.price_view()}\n"
            f"📦 <b>В наличии:</b> {item.stock}\n",
            reply_markup=builder.as_markup(),
            parse_mode="HTML"
        )
        return

    photo = FSInputFile(item.image.path)

    caption = (
        f"📦 {item.title}\n\n"
        f"📝 <b>Описание:</b>\n{item.description[:500]}...\n"
        f"💵 <b>Цена:</b> {item.price_view()}\n"
        f"📦 <b>В наличии:</b> {item.stock}\n"
    )

    await callback.message.answer_photo(
        photo=photo,
        caption=caption,
        reply_markup=builder.as_markup(),
        parse_mode="HTML"
    )


@dp.callback_query(F.data.startswith("add_to_cart_"))
async def add_to_cart(callback: CallbackQuery, state: FSMContext):
    item_id = int(callback.data.split("_")[-1])
    await state.update_data(item_id=item_id)
    await callback.message.answer("Введите количество:")
    await state.set_state(CartStates.waiting_for_quantity)


async def render_cart(user_id: int, target, edit_mode=False):
    """Вспомогательная функция для отображения корзины"""
    cart_items = await TelegramUser.async_get_cart_items(user_id=user_id)

    if not cart_items:
        await target.answer("Ваша корзина пуста.", reply_markup=KB.back)
        return

    cart_text = "\n".join(
        f"{item.quantity} x {item.item.title} — {intcomma(item.total_price())} ₽"
        for item in cart_items
    )
    total_price = sum(item.total_price() for item in cart_items)

    builder = InlineKeyboardBuilder()
    for item in cart_items:
        builder.button(
            text=f"❌ Удалить {item.item.title}",
            callback_data=f"remove_item_{item.id}"
        )
    builder.button(text="Оформить заказ", callback_data="checkout")
    builder.button(text="🔙 Назад", callback_data="menu")
    builder.adjust(1)

    message_text = f"🛒 Ваша корзина:\n\n{cart_text}\n\n💰 Общая сумма: {intcomma(total_price)} ₽"

    if edit_mode:
        await target.edit_text(
            text=message_text,
            reply_markup=builder.as_markup()
        )
    else:
        await target.answer(
            text=message_text,
            reply_markup=builder.as_markup()
        )


@dp.message(F.text, StateFilter(CartStates.waiting_for_quantity))
async def process_quantity(message: Message, state: FSMContext):
    data = await state.get_data()
    item_id = data.get("item_id")
    item = await sync_to_async(Item.objects.get)(id=item_id)

    try:
        quantity = int(message.text)
        if quantity <= 0 or quantity > item.stock:
            await message.answer(f"Введите число от 1 до {item.stock}.")
            return
    except ValueError:
        await message.answer("Пожалуйста, введите число.")
        return

    user_id = message.from_user.id
    user = await sync_to_async(TelegramUser.objects.get)(user_id=user_id)
    shopping_cart, _ = await sync_to_async(ShoppingCart.objects.get_or_create)(user=user)

    cart_item, created = await sync_to_async(CartItem.objects.get_or_create)(
        shopping_cart=shopping_cart,
        item=item,
        defaults={"quantity": quantity}
    )

    if not created:
        cart_item.quantity += quantity
        await sync_to_async(cart_item.save)()

    item.stock -= quantity
    await sync_to_async(item.save)()

    await message.answer(f"Товар {item.title} добавлен в корзину!")
    await state.clear()

    await render_cart(
        user_id=message.from_user.id,
        target=message,
        edit_mode=False
    )


@dp.callback_query(F.data == "shopping_cart")
async def open_shopping_cart(callback: CallbackQuery):
    await bot.answer_callback_query(callback.id)

    await render_cart(
        user_id=callback.from_user.id,
        target=callback.message,
        edit_mode=True
    )


@dp.callback_query(F.data.startswith("remove_item_"))
async def remove_item_from_cart(callback: CallbackQuery):
    item_id = int(callback.data.split("_")[-1])

    cart_item = await sync_to_async(CartItem.objects.select_related('item').get)(id=item_id)

    item = cart_item.item
    item.stock += cart_item.quantity
    await sync_to_async(item.save)()
    await sync_to_async(cart_item.delete)()

    await callback.answer(f"Товар {item.title} удален из корзины!")

    await render_cart(
        user_id=callback.from_user.id,
        target=callback.message,
        edit_mode=True
    )


@dp.callback_query(F.data == "checkout")
async def checkout(callback: CallbackQuery, state: FSMContext):
    await callback.answer()

    await callback.message.answer("📦 Пожалуйста, введите адрес доставки:")

    await state.set_state(CheckoutStates.waiting_for_address)


@dp.message(StateFilter(CheckoutStates.waiting_for_address))
async def process_address(message: Message, state: FSMContext):
    user_id = message.from_user.id

    address = message.text
    await state.update_data(address=address)

    total_amount = await get_cart_total(user_id)
    total_in_kopecks = int(total_amount * 100)

    if total_in_kopecks < 1 or total_in_kopecks > 100_000_000:
        await message.answer("Сумма заказа должна быть от 0.01 до 1 000 000 рублей.")
        return

    builder = InlineKeyboardBuilder()
    builder.button(text="💳 Оплатить", pay=True)

    await message.answer_invoice(
        title="Оплата заказа",
        description="Оплата заказа через ЮКассу",
        payload="order_payload",
        provider_token=config.PROVIDER_TOKEN,
        currency="RUB",
        prices=[
            LabeledPrice(label="Заказ", amount=total_in_kopecks)
        ],
        start_parameter="create_invoice",
        reply_markup=builder.as_markup()
    )

    await state.clear()


@dp.pre_checkout_query()
async def process_pre_checkout_query(pre_checkout_query: PreCheckoutQuery):
    await bot.answer_pre_checkout_query(pre_checkout_query.id, ok=True)


@dp.message(F.successful_payment)
async def process_successful_payment(message: Message, state: FSMContext):
    user_id = message.from_user.id
    data = await state.get_data()
    address = data.get("address")

    user = await sync_to_async(TelegramUser.objects.get)(user_id=user_id)
    cart = await sync_to_async(ShoppingCart.objects.get)(user=user)
    cart_items = await sync_to_async(list)(cart.items.all())

    items = [f"{item.quantity} x {item.item.title}" for item in cart_items]

    await sync_to_async(save_order_to_excel)(
        user_id=user_id,
        address=address,
        total_amount=message.successful_payment.total_amount // 100,
        items=items
    )

    await sync_to_async(cart.items.clear)()

    await message.answer(
        "✅ Оплата прошла успешно! Ваш заказ будет доставлен по адресу:\n"
        f"📦 {address}\n\n"
        f"💰 Сумма заказа: {message.successful_payment.total_amount // 100} руб."
    )


@dp.callback_query(F.data.startswith("prev_page_"))
async def prev_page(callback: CallbackQuery, state: FSMContext):
    page = int(callback.data.split("_")[-1])
    await state.update_data(page=page - 1)
    await display_category_items(callback, state)


@dp.callback_query(F.data.startswith("next_page_"))
async def next_page(callback: CallbackQuery, state: FSMContext):
    page = int(callback.data.split("_")[-1])
    await state.update_data(page=page + 1)
    await display_category_items(callback, state)


@dp.callback_query(F.data == "faq")
async def show_faq(callback: CallbackQuery):
    await callback.message.answer(
        text="❓ <b>Часто задаваемые вопросы</b>\n\n"
             "Вы можете найти ответы на популярные вопросы, используя инлайн-режим.\n"
             "Просто начните вводить @post_autoforward_bot и ваш вопрос.\n\n"
             "Например: <code>@post_autoforward_bot оплата</code>",
        parse_mode="HTML"
    )


@dp.inline_query()
async def inline_faq(inline_query: InlineQuery):
    query = inline_query.query.lower()
    results = []

    for faq in faq_data:
        if query in faq["question"].lower():
            results.append(
                InlineQueryResultArticle(
                    id=str(faq_data.index(faq)),
                    title=faq["question"],
                    input_message_content=InputTextMessageContent(
                        message_text=f"❓ <b>{faq['question']}</b>\n\n{faq['answer']}",
                        parse_mode="HTML"
                    ),
                    description=faq["answer"],
                )
            )

    await inline_query.answer(results, cache_time=1)


async def main():
    await bot(DeleteWebhook(drop_pending_updates=True))
    await dp.start_polling(bot)


if __name__ == '__main__':
    logging.basicConfig(
        level=logging.INFO,
        format=('%(asctime)s, '
                '%(levelname)s, '
                '%(funcName)s, '
                '%(message)s'
                ),
        encoding='UTF-8',
        handlers=[logging.FileHandler(__file__ + '.log'),
                  logging.StreamHandler(sys.stdout)]
    )
    asyncio.run(main())
